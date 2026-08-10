import os
import openpyxl

folder = r"E:\earthyelectronics\backend\all products files"
led_xlsx = os.path.join(folder, "LED UPLOADING.xlsx")
disp_xlsx = os.path.join(folder, "water dispenser.xlsx")

print("==================================================")
print("📄 OPENPYXL EXCEL COLUMNS & IMAGE URLS INSPECTION")
print("==================================================")

if os.path.exists(led_xlsx):
    try:
        wb = openpyxl.load_workbook(led_xlsx)
        sheet = wb.active
        print(f"\n📺 LED UPLOADING.xlsx (Sheet: {sheet.title}, Max Row: {sheet.max_row}):")
        for row in list(sheet.iter_rows(values_only=True))[:10]:
            print("   ", [str(c)[:50] for c in row if c is not None])
    except Exception as e:
        print("Error reading LED xlsx:", e)

if os.path.exists(disp_xlsx):
    try:
        wb = openpyxl.load_workbook(disp_xlsx)
        sheet = wb.active
        print(f"\n💧 water dispenser.xlsx (Sheet: {sheet.title}, Max Row: {sheet.max_row}):")
        for row in list(sheet.iter_rows(values_only=True))[:10]:
            print("   ", [str(c)[:50] for c in row if c is not None])
    except Exception as e:
        print("Error reading Dispenser xlsx:", e)

print("==================================================")
